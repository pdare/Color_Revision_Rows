import openpyxl
from openpyxl.styles import PatternFill
from enum import Enum

color_1 = 'F4B084'
color_2 = 'FFE699'
current_color = color_1
current_name = ''
last_name = ''


def switch_color(curColor):
    if curColor == color_1:
        return color_2
    else:
        return color_1

wb = openpyxl.load_workbook('current_revisions.xlsx')
fs = wb.active
fs_count_row = fs.max_row 
fs_count_col = fs.max_column

LAST_NAME_F = 1
FIRST_NAME_F = 2
OLD_VAL = 7
NEW_VAL = 8
FIELD_REVISED = 6

removed_row_count = 0

#fs.cell(row = 1, column = 1).fill = PatternFill(fgColor='008800', fill_type = 'solid')
print(fs.cell(row = 1, column = 1).value)

for row in range(2, fs_count_row+1):
    if fs.cell(row = row, column=NEW_VAL).value is None:
        fs.delete_rows(row)
        print("row " + str(row) + " was deleted for blank new value")
        removed_row_count += 1
    elif fs.cell(row = row, column=FIELD_REVISED).value == "Parent1 Middle Name" or fs.cell(row = row, column=FIELD_REVISED).value == "Parent2 Middle Name":
        print("row " + str(row) + " was deleted for " + str(fs.cell(row = row, column=FIELD_REVISED).value))
        fs.delete_rows(row)
        removed_row_count += 1
    elif fs.cell(row = row, column=OLD_VAL).value == fs.cell(row = row, column=NEW_VAL).value:
        print("row " + str(row) + " was deleted for old and new values being the same: " + fs.cell(row = row, column=OLD_VAL).value + " : " + fs.cell(row = row, column=NEW_VAL).value)
        fs.delete_rows(row)
        removed_row_count += 1
    elif "Name" in fs.cell(row=row, column=FIELD_REVISED).value:
        string_chck = str(fs.cell(row=row, column=NEW_VAL).value)
        length = len(string_chck) - 1
        if str(fs.cell(row=row, column=NEW_VAL).value)[length:] == " ":
            print("row " + str(row) + " deleted for name with end space")
            removed_row_count += 1
        

fs_count_row = fs.max_row
fs_count_col = fs.max_column

current_name = fs.cell(row = 2, column = LAST_NAME_F).value + fs.cell(row = 2, column = FIRST_NAME_F).value
last_name = fs.cell(row = 2, column = LAST_NAME_F).value + fs.cell(row = 2, column = FIRST_NAME_F).value

for row in range(2,fs_count_row+1):
    current_name = fs.cell(row = row, column = LAST_NAME_F).value + fs.cell(row = row, column = FIRST_NAME_F).value
    if current_name == last_name:
        for column in range(1,fs_count_col+1):
            fs.cell(row = row, column = column).fill = PatternFill(fgColor=current_color, fill_type = 'solid')
        last_name = current_name
    else:
        current_color = switch_color(current_color)
        for column in range(1,fs_count_col+1):
            fs.cell(row = row, column = column).fill = PatternFill(fgColor=current_color, fill_type = 'solid')
        last_name = current_name
wb.save('current_revisions_colored.xlsx')
print(str(removed_row_count) + " rows removed")
print("done")

